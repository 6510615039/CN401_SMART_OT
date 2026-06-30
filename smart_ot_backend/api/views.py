import re
import unicodedata
from django.utils import timezone
from django.db.models import Q, Case, When, Value, IntegerField
from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
import openpyxl
from .models import (
    User, Department, OTRequest, Holiday,
    SystemSettings, TimeLog, ImportHistory, AuditLog, OTDeadline,
    Notification, NoOTDeclaration, DepartmentMonthlyBudget,
)
from .serializers import (
    UserSerializer, UserCreateSerializer, DepartmentSerializer,
    OTRequestSerializer, HolidaySerializer, SystemSettingsSerializer,
    TimeLogSerializer, ImportHistorySerializer, AuditLogSerializer,
    NotificationSerializer,
)


def get_effective_role(user, request):
    """ดึง role ที่ใช้งานจริง — รองรับ X-Acting-Role header สำหรับ multi-role users
    ถ้า header ส่งมาและ user มี role นั้นใน available_roles → ใช้ role นั้น
    มิฉะนั้นใช้ user.role ปกติ
    """
    acting_as = request.META.get('HTTP_X_ACTING_ROLE', '').strip()
    if acting_as:
        available = getattr(user, 'available_roles', None) or []
        if acting_as in available:
            return acting_as
    return user.role


def _send_checker_notification(ot_list, note, sender):
    """ส่งอีเมลสรุปคำร้อง OT ที่ส่งต่อให้ checker ทุกคน (fail-silent)"""
    from django.core.mail import send_mail
    from django.conf import settings as djsettings
    try:
        checkers = User.objects.filter(role='checker', is_active=True)
        recipients = [
            u.notify_email or u.email
            for u in checkers
            if (u.notify_email or u.email)
        ]
        if not recipients:
            return
        dept_name = ot_list[0].department.name if ot_list else ''
        total_amount = sum(float(ot.amount) for ot in ot_list)
        total_hours  = sum(float(ot.ot_hours) for ot in ot_list)
        subject = f'[SMART OT] คำร้อง OT {len(ot_list)} รายการ รอตรวจสอบ — {dept_name}'
        lines = [
            'เรียน ผู้ตรวจสอบ',
            '',
            f'ตัวแทนแผนก {dept_name} ({sender.get_full_name()}) ส่งต่อคำร้อง OT จำนวน {len(ot_list)} รายการ:',
            '',
        ]
        for ot in ot_list:
            lines.append(f'  • {ot.staff.get_full_name()} — {ot.work_date}  {float(ot.ot_hours):.1f} ชม. / {float(ot.amount):,.0f} บาท')
        lines += [
            '',
            f'รวม: {total_hours:.1f} ชั่วโมง  —  {total_amount:,.0f} บาท',
        ]
        if note:
            lines += ['', f'หมายเหตุจากตัวแทนแผนก: {note}']
        lines += ['', 'กรุณาเข้าสู่ระบบ SMART OT เพื่อดำเนินการ']
        send_mail(subject, '\n'.join(lines),
                  djsettings.DEFAULT_FROM_EMAIL, recipients,
                  fail_silently=True)
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f'send_checker_notification: {e}')


def _send_email(user, subject: str, body: str):
    """ส่งเมลให้ user คนเดียว (fail-silent) ใช้ notify_email ก่อน ถ้าไม่มีใช้ email"""
    try:
        from django.core.mail import send_mail
        from django.conf import settings as djsettings
        email = getattr(user, 'notify_email', '') or getattr(user, 'email', '')
        if email:
            send_mail(subject, body, djsettings.DEFAULT_FROM_EMAIL, [email], fail_silently=True)
    except Exception:
        pass


def _push_ws(user_id: int, notif_data: dict):
    """Push notification ไปยัง WebSocket channel layer ของ user (fail-silent)"""
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        layer = get_channel_layer()
        if layer:
            async_to_sync(layer.group_send)(
                f'user_{user_id}',
                {'type': 'notification.send', 'data': notif_data},
            )
    except Exception:
        pass


def _notify_ot(ot, notif_type, recipients, message):
    """สร้าง Notification + ส่งอีเมล + push WebSocket (fail-silent) ให้ผู้รับที่ระบุ"""
    for user in recipients:
        notif = Notification.objects.create(
            recipient=user,
            message=message,
            notif_type=notif_type,
            ot_request=ot,
        )
        # Push real-time ผ่าน WebSocket
        _push_ws(user.id, {
            'id': notif.id,
            'message': notif.message,
            'notif_type': notif.notif_type,
            'ot_request': ot.id if ot else None,
            'ot_request_date': str(ot.work_date) if ot else None,
            'is_read': False,
            'created_at': notif.created_at.isoformat(),
        })
    # ส่งอีเมลทีละคน (ใช้ notify_email ก่อน ถ้าไม่มีใช้ email)
    body = f'{message}\n\nวันที่ OT: {ot.work_date}  {float(ot.ot_hours):.1f} ชม.  {float(ot.amount):,.0f} บาท\n\nกรุณาเข้าสู่ระบบ SMART OT เพื่อดำเนินการ'
    for user in recipients:
        _send_email(user, f'[SMART OT] {message[:60]}', body)


def log_action(user, action, model_name='', object_id='', detail='', request=None):
    ip = None
    if request:
        x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
        ip = x_forwarded.split(',')[0] if x_forwarded else request.META.get('REMOTE_ADDR')
    AuditLog.objects.create(
        user=user, action=action,
        model_name=model_name, object_id=str(object_id),
        detail=detail, ip_address=ip
    )


# ─── Auth ────────────────────────────────────────────────────────────────────

def _auto_create_user_from_tu(tu_data: dict) -> 'User | None':
    """
    สร้างหรืออัปเดต User จากข้อมูล TU Auth
    ลำดับการค้นหา:
      1. หาจาก email (ตรงกับ user ที่ import จากไฟล์ 07 ซึ่งมี email_tu)
      2. หาจาก TU username
      3. ถ้าไม่เจอ: สร้างใหม่ด้วย role='staff'
    คง role เดิมเสมอ ไม่ overwrite
    """
    username = tu_data.get('username', '').strip()
    if not username:
        return None

    dept_name = tu_data.get('dept_name', '')
    dept = Department.objects.get_or_create(
        name=dept_name, defaults={'code': dept_name[:15].upper().replace(' ', '_')}
    )[0] if dept_name else None
    defaults = {
        'first_name': tu_data.get('first_name', ''),
        'last_name':  tu_data.get('last_name', ''),
        'email':      tu_data.get('email', ''),
        'is_active':  True,
    }
    if dept:
        defaults['department'] = dept

    try:
        email = tu_data.get('email', '').strip()

        # 1) หาจาก email ก่อน (user ที่ import จากไฟล์ 07 มี email_tu ตรงกัน)
        user = None
        if email:
            user = User.objects.filter(email__iexact=email).first()

        # 2) หาจาก TU username
        if not user:
            user = User.objects.filter(username=username).first()

        # 3) หาจาก employee_id (กรณี username ในระบบเป็น employee_id เช่น 0006)
        if not user:
            user = User.objects.filter(employee_id=username).first()

        if user:
            # อัปเดตชื่อ/email/dept แต่ไม่เปลี่ยน role
            for k, v in defaults.items():
                setattr(user, k, v)
            user.save()
        else:
            # สร้างใหม่ด้วย role='staff'
            user = User(username=username, role='staff', **defaults)
            user.set_unusable_password()
            user.save()
        return user
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f'auto_create_user_from_tu: {e}')
        return None


@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    """
    Login flow:
    1. TU API Auth (ถ้าเปิดใช้งาน) → auto-create/sync user → JWT
    2. Local Django auth (fallback หรือบัญชี admin)

    รองรับทั้ง:
      - อีเมล TU: s6512345678@dome.tu.ac.th / sariya@tu.ac.th
      - username: somchai / s6512345678
    """
    from .tu_api_service import verify_tu_credentials

    username_input = request.data.get('username', '').strip()
    password       = request.data.get('password', '').strip()

    if not username_input or not password:
        return Response({'error': 'กรุณากรอกอีเมลและรหัสผ่าน'}, status=400)

    # ตัด @domain ออกเพื่อใช้เป็น TU username
    tu_username = username_input.split('@')[0] if '@' in username_input else username_input

    user = None

    # ── ขั้นตอน 1: TU API Auth ─────────────────────────────────────────
    tu_data = verify_tu_credentials(tu_username, password)
    if tu_data:
        user = _auto_create_user_from_tu(tu_data)
        if not user:
            return Response({'error': 'ไม่สามารถสร้างบัญชีในระบบได้ กรุณาติดต่อผู้ดูแลระบบ'}, status=500)

    # ── ขั้นตอน 2: Local Auth (fallback) ──────────────────────────────
    if not user:
        if '@' in username_input:
            # ค้นจาก email field ก่อน
            try:
                found = User.objects.get(email__iexact=username_input)
                user = authenticate(request, username=found.username, password=password)
            except User.DoesNotExist:
                pass
            # ลอง local username (ส่วนก่อน @)
            if not user:
                user = authenticate(request, username=tu_username, password=password)
        else:
            user = authenticate(request, username=username_input, password=password)
            if not user:
                try:
                    found = User.objects.get(employee_id=username_input)
                    user = authenticate(request, username=found.username, password=password)
                except User.DoesNotExist:
                    pass

    if not user:
        return Response({'error': 'อีเมลหรือรหัสผ่านไม่ถูกต้อง'}, status=401)

    if not user.is_active:
        return Response({'error': 'บัญชีนี้ถูกระงับการใช้งาน'}, status=403)

    refresh = RefreshToken.for_user(user)
    log_action(user, 'เข้าสู่ระบบ', request=request)

    return Response({
        'access':  str(refresh.access_token),
        'refresh': str(refresh),
        'user': UserSerializer(user).data,
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def refresh_token_view(request):
    from rest_framework_simplejwt.views import TokenRefreshView
    return TokenRefreshView.as_view()(request._request)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me_view(request):
    return Response(UserSerializer(request.user).data)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def me_update_view(request):
    user = request.user
    profile_image = request.data.get('profile_image', '')
    if profile_image:
        # รับ base64 data URL เช่น "data:image/jpeg;base64,..."
        # จำกัดขนาดไม่เกิน 5MB (base64 ~4/3 ของขนาดจริง)
        if len(profile_image) > 7 * 1024 * 1024:
            return Response({'error': 'image too large'}, status=400)
        user.profile_image = profile_image
        user.save(update_fields=['profile_image'])
    return Response({'status': 'ok', 'profile_image': user.profile_image})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    log_action(request.user, 'ออกจากระบบ', request=request)
    return Response({'message': 'ออกจากระบบสำเร็จ'})


# ─── Users ───────────────────────────────────────────────────────────────────

class UserViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    queryset = User.objects.all().annotate(
        dept_order=Case(
            When(department__name='ผู้บริหาร', then=Value(0)),
            When(department__name__startswith='งานทะเบียนนักศึกษา', then=Value(1)),
            When(department__name__startswith='งานเทคโนโลยี', then=Value(2)),
            When(department__name__startswith='งานยุทธศาสตร์', then=Value(3)),
            default=Value(9),
            output_field=IntegerField(),
        )
    ).order_by('dept_order', 'department__name', 'employee_id')

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        role = self.request.query_params.get('role')
        dept = self.request.query_params.get('department')
        q    = self.request.query_params.get('q')
        if role: qs = qs.filter(role=role)
        if dept: qs = qs.filter(department_id=dept)
        if q:    qs = qs.filter(Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(username__icontains=q))
        return qs

    def perform_create(self, serializer):
        user = serializer.save()
        log_action(self.request.user, f'เพิ่มผู้ใช้ {user.get_full_name()}', 'User', user.id, request=self.request)

    def perform_destroy(self, instance):
        name = instance.get_full_name()
        instance.is_active = False
        instance.save()
        log_action(self.request.user, f'ระงับผู้ใช้ {name}', 'User', instance.id, request=self.request)


# ─── Departments ─────────────────────────────────────────────────────────────

class DepartmentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = DepartmentSerializer
    queryset = Department.objects.all().order_by('name')


# ─── OT Requests ─────────────────────────────────────────────────────────────

class OTRequestViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = OTRequestSerializer

    def get_queryset(self):
        user = self.request.user
        effective_role = get_effective_role(user, self.request)
        qs = OTRequest.objects.select_related('staff', 'department')

        # กรองตาม effective_role (รองรับ multi-role users ที่ส่ง X-Acting-Role header)
        if effective_role == 'staff':
            qs = qs.filter(staff=user)
        elif effective_role == 'depthead':
            qs = qs.filter(staff__department=user.department)
        elif effective_role == 'deptrep':
            if user.department:
                # ใช้ OR: ตรงกับ OTRequest.department หรือ staff.department
                qs = qs.filter(
                    Q(department=user.department) | Q(staff__department=user.department)
                ).distinct()
            # ถ้า deptrep ไม่มีแผนก: เห็นทั้งหมดในสถานะที่เกี่ยวข้อง
            # default (ไม่มี ?status): แสดงเฉพาะที่รอส่งต่อ
            if not self.request.query_params.get('status'):
                qs = qs.filter(status='head_approved')
        elif effective_role == 'checker':
            qs = qs.filter(status__in=['rep_forwarded', 'checker_approved', 'checker_rejected', 'completed'])
        # admin/executive เห็นทั้งหมด

        # filter params
        status_p = self.request.query_params.get('status')
        dept_p   = self.request.query_params.get('department')
        month_p  = self.request.query_params.get('month')  # YYYY-MM
        if status_p: qs = qs.filter(status=status_p)
        status_in_p = self.request.query_params.get('status_in')
        if status_in_p: qs = qs.filter(status__in=status_in_p.split(','))
        if dept_p:   qs = qs.filter(department_id=dept_p)
        # checker ประวัติของตัวเอง
        if self.request.query_params.get('approved_by_me'):
            qs = qs.filter(checker_approved_by=user)
        if month_p:
            year, month = month_p.split('-')
            qs = qs.filter(work_date__year=year, work_date__month=month)

        return qs.order_by('-work_date')

    def create(self, request, *args, **kwargs):
        # ตรวจ deadline ก่อน create — ถ้าเลยกำหนดปฏิเสธทันที
        import datetime as _dt
        work_date_str = request.data.get('work_date', '')
        if work_date_str:
            try:
                work_date_obj = _dt.date.fromisoformat(work_date_str)
                # แปลงเป็น thai_month (ปี พ.ศ.)
                thai_year = work_date_obj.year + 543
                thai_month = f'{thai_year}-{work_date_obj.month:02d}'
                err = _check_ot_deadline(thai_month)
                if err:
                    return Response({'error': err}, status=status.HTTP_400_BAD_REQUEST)
            except ValueError:
                pass
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        # คำนวณ day_type อัตโนมัติ
        work_date = serializer.validated_data.get('work_date')
        is_holiday = Holiday.objects.filter(date=work_date).exists()
        is_weekend = work_date.weekday() >= 5
        day_type = 'holiday' if (is_holiday or is_weekend) else 'weekday'

        # คำนวณค่าตอบแทน: วันธรรมดา 60 บาท/ชม. max 4ชม, วันหยุด 70 บาท/ชม. max 7ชม
        max_hours = 7 if day_type == 'holiday' else 4
        # ot_hours เป็น read_only ใน serializer → ต้องอ่านจาก request.data โดยตรง
        ot_hours_raw = float(self.request.data.get('ot_hours', 0))
        ot_hours = min(ot_hours_raw, max_hours)  # cap ที่ ceiling เสมอ
        hourly_rate = 70 if day_type == 'holiday' else 60
        amount = ot_hours * hourly_rate

        # ถ้า user ไม่มีแผนก ให้ใช้แผนก default หรือสร้างใหม่
        dept = self.request.user.department
        if dept is None:
            from .models import Department
            dept, _ = Department.objects.get_or_create(
                name='ไม่ระบุแผนก', defaults={'code': 'NONE'}
            )

        ot = serializer.save(
            staff=self.request.user,
            department=dept,
            day_type=day_type,
            ot_hours=ot_hours,
            rate_per_hour=hourly_rate,
            amount=amount,
            status='submitted',
        )
        log_action(self.request.user, f'ยื่นคำร้อง OT วันที่ {ot.work_date}', 'OTRequest', ot.id, request=self.request)

        # แจ้งหัวหน้าแผนก
        staff_name = self.request.user.get_full_name() or self.request.user.username
        deptheads = list(User.objects.filter(role='depthead', department=dept, is_active=True))
        if deptheads:
            _notify_ot(
                ot, 'ot_submitted', deptheads,
                f'{staff_name} ยื่นคำร้อง OT วันที่ {ot.work_date} จำนวน {float(ot_hours):.0f} ชม. ({float(amount):,.0f} บาท) รอการอนุมัติ',
            )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        ot = self.get_object()
        user = request.user
        effective_role = get_effective_role(user, request)
        note = request.data.get('note', '')

        head_name    = user.get_full_name() or user.username
        staff_name_a = ot.staff.get_full_name() or ot.staff.username

        if effective_role == 'depthead' and ot.status == 'submitted':
            # ตรวจเพดานงบก่อน approve
            work_dt = ot.work_date
            _year, _mon = work_dt.year, work_dt.month
            try:
                _mb = DepartmentMonthlyBudget.objects.get(department=ot.department, year=_year, month=_mon)
                _monthly_budget = float(_mb.budget)
            except DepartmentMonthlyBudget.DoesNotExist:
                _monthly_budget = 0
            if _monthly_budget > 0:
                _used = sum(
                    _calc_ot_amount(r.ot_hours, r.day_type)
                    for r in OTRequest.objects.filter(
                        department=ot.department,
                        work_date__year=_year, work_date__month=_mon,
                        status__in=['head_approved', 'rep_forwarded', 'checker_approved', 'completed'],
                    )
                )
                _this = _calc_ot_amount(ot.ot_hours, ot.day_type)
                if _used + _this > _monthly_budget:
                    _over = _used + _this - _monthly_budget
                    return Response(
                        {'error': f'งบประมาณ OT แผนก {ot.department.name} เดือนนี้เต็มแล้ว '
                                  f'(ใช้ไป {_used:,.0f} / {_monthly_budget:,.0f} บาท — '
                                  f'คำร้องนี้ต้องการ {_this:,.0f} บาท เกิน {_over:,.0f} บาท)'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            ot.status = 'head_approved'
            ot.head_approved_by = user
            ot.head_approved_at = timezone.now()
            ot.head_note = note
            ot.save()
            log_action(user, f'อนุมัติคำร้อง OT #{ot.id}', 'OTRequest', ot.id, request=request)
            # ไม่แจ้งพนักงานตอนนี้ — พนักงานจะรู้เมื่อ checker อนุมัติ/ตีกลับขั้นสุดท้ายเท่านั้น
            # แจ้ง deptrep ในแผนกเดียวกัน (ใช้ type แยกจาก staff)
            deptreps = list(User.objects.filter(role='deptrep', department=ot.department, is_active=True))
            if deptreps:
                _notify_ot(ot, 'ot_rep_action_needed', deptreps,
                    f'{staff_name_a} ได้รับการอนุมัติจากหัวหน้างานแล้ว รอตัวแทนฝ่ายส่งต่อ')
            return Response({'message': 'อนุมัติสำเร็จ'})

        elif effective_role == 'deptrep' and ot.status == 'head_approved':
            ot.status = 'rep_forwarded'
            ot.rep_forwarded_by = user
            ot.rep_forwarded_at = timezone.now()
            ot.rep_note = note
            ot.save()
            log_action(user, f'ส่งต่อคำร้อง OT #{ot.id}', 'OTRequest', ot.id, request=request)
            # แจ้งพนักงาน
            _notify_ot(ot, 'ot_rep_forwarded', [ot.staff],
                f'ตัวแทนฝ่ายส่งต่อคำร้อง OT วันที่ {ot.work_date} ของคุณให้ผู้ตรวจสอบแล้ว')
            return Response({'message': 'ส่งต่อสำเร็จ'})

        elif effective_role == 'checker' and ot.status == 'rep_forwarded':
            # ตรวจสอบเพดานงบประมาณรายเดือนของแผนก
            import datetime as _dt
            work_dt = ot.work_date
            year, mon = work_dt.year, work_dt.month
            try:
                mb = DepartmentMonthlyBudget.objects.get(
                    department=ot.staff.department, year=year, month=mon
                )
                monthly_budget = float(mb.budget)
            except DepartmentMonthlyBudget.DoesNotExist:
                monthly_budget = 0

            budget_warning = None
            if monthly_budget > 0:
                already_used = sum(
                    _calc_ot_amount(r.ot_hours, r.day_type)
                    for r in OTRequest.objects.filter(
                        staff__department=ot.staff.department,
                        work_date__year=year,
                        work_date__month=mon,
                        status__in=APPROVED_STATUSES,
                    )
                )
                this_amount = _calc_ot_amount(ot.ot_hours, ot.day_type)
                new_total = already_used + this_amount
                if new_total > monthly_budget:
                    over_by = new_total - monthly_budget
                    pct = round(new_total / monthly_budget * 100)
                    budget_warning = (
                        f'⚠ ยอด OT แผนก {ot.staff.department.name} จะเกินเพดานงบ '
                        f'{monthly_budget:,.0f} บาท ({pct}% — เกิน {over_by:,.0f} บาท)'
                    )

            ot.status = 'checker_approved'
            ot.checker_approved_by = user
            ot.checker_approved_at = timezone.now()
            ot.checker_note = note
            ot.save()
            log_action(user, f'ผู้ตรวจสอบอนุมัติคำร้อง OT #{ot.id}', 'OTRequest', ot.id, request=request)
            # notification รายบุคคลสำหรับกรณีอนุมัติเดี่ยว (bulk ใช้ bulk_approve_view แทน)
            checker_name = user.get_full_name() or user.username
            _notify_ot(ot, 'ot_checker_approved', [ot.staff],
                f'ผู้ตรวจสอบ ({checker_name}) อนุมัติคำร้อง OT วันที่ {ot.work_date} ของคุณแล้ว — {float(ot.amount):,.0f} บาท')
            resp = {'message': 'อนุมัติสำเร็จ'}
            if budget_warning:
                resp['budget_warning'] = budget_warning
            return Response(resp)

        return Response({'error': f'ไม่สามารถดำเนินการได้ (role={effective_role}, status={ot.status})'}, status=400)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        ot = self.get_object()
        user = request.user
        effective_role = get_effective_role(user, request)
        note = request.data.get('note', '')

        actor_name   = user.get_full_name() or user.username
        staff_name_r = ot.staff.get_full_name() or ot.staff.username

        if effective_role == 'depthead' and ot.status == 'submitted':
            ot.status = 'head_rejected'
            ot.head_note = note
            ot.save()
            log_action(user, f'ตีกลับคำร้อง OT #{ot.id}', 'OTRequest', ot.id, request=request)
            # แจ้งพนักงาน
            reason_text = f' — เหตุผล: {note}' if note else ''
            _notify_ot(ot, 'ot_head_rejected', [ot.staff],
                f'หัวหน้างาน ({actor_name}) ตีกลับคำร้อง OT วันที่ {ot.work_date} ของคุณ{reason_text}')
            return Response({'message': 'ตีกลับสำเร็จ'})

        elif effective_role == 'checker' and ot.status == 'rep_forwarded':
            ot.status = 'checker_rejected'
            ot.checker_note = note
            ot.save()
            log_action(user, f'ผู้ตรวจสอบตีกลับคำร้อง OT #{ot.id}', 'OTRequest', ot.id, request=request)
            # แจ้งพนักงานและหัวหน้าแผนก
            reason_text = f' — เหตุผล: {note}' if note else ''
            _notify_ot(ot, 'ot_checker_rejected', [ot.staff],
                f'ผู้ตรวจสอบ ({actor_name}) ตีกลับคำร้อง OT วันที่ {ot.work_date} ของคุณ{reason_text}')
            deptheads_r = list(User.objects.filter(role='depthead', department=ot.department, is_active=True))
            if deptheads_r:
                _notify_ot(ot, 'ot_checker_rejected', deptheads_r,
                    f'คำร้อง OT ของ {staff_name_r} วันที่ {ot.work_date} ถูกผู้ตรวจสอบตีกลับ{reason_text}')
            # แจ้ง deptrep ในแผนกเดียวกัน
            deptreps_r = list(User.objects.filter(role='deptrep', department=ot.department, is_active=True))
            if deptreps_r:
                _notify_ot(ot, 'ot_checker_rejected', deptreps_r,
                    f'ผู้ตรวจสอบ ({actor_name}) ตีกลับคำร้อง OT ของ {staff_name_r} วันที่ {ot.work_date}{reason_text}')
            return Response({'message': 'ตีกลับสำเร็จ'})

        return Response({'error': 'ไม่สามารถดำเนินการได้ในสถานะนี้'}, status=400)

    @action(detail=True, methods=['post'])
    def resubmit(self, request, pk=None):
        """พนักงานยื่นคำร้องซ้ำหลังถูกตีกลับ — อัปเดต request เดิม ไม่สร้างใหม่"""
        ot = self.get_object()
        if ot.status not in ('head_rejected', 'checker_rejected'):
            return Response({'error': 'คำร้องนี้ไม่ได้ถูกตีกลับ'}, status=400)
        if ot.staff != request.user:
            return Response({'error': 'ไม่มีสิทธิ์แก้ไขคำร้องนี้'}, status=403)

        # คำนวณใหม่พร้อม cap
        max_hours = 7 if ot.day_type == 'holiday' else 4
        ot_hours_raw = float(request.data.get('ot_hours', ot.ot_hours))
        ot_hours = min(ot_hours_raw, max_hours)
        hourly_rate = 70 if ot.day_type == 'holiday' else 60

        ot.ot_hours = ot_hours
        ot.amount = ot_hours * hourly_rate
        ot.work_detail = request.data.get('work_detail', ot.work_detail)
        ot.start_time = request.data.get('start_time', ot.start_time)
        ot.end_time = request.data.get('end_time', ot.end_time)
        # reset สถานะ + ล้าง note เดิม
        ot.status = 'submitted'
        ot.head_note = ''
        ot.checker_note = ''
        ot.save()

        log_action(request.user, f'ยื่นคำร้อง OT #{ot.id} ซ้ำหลังถูกตีกลับ วันที่ {ot.work_date}', 'OTRequest', ot.id, request=request)

        # แจ้งหัวหน้าแผนกว่ามีการยื่นซ้ำ
        staff_name_rs = request.user.get_full_name() or request.user.username
        deptheads_rs = list(User.objects.filter(role='depthead', department=ot.department, is_active=True))
        if deptheads_rs:
            _notify_ot(ot, 'ot_submitted', deptheads_rs,
                f'{staff_name_rs} ยื่นคำร้อง OT วันที่ {ot.work_date} ซ้ำหลังถูกตีกลับ จำนวน {float(ot.ot_hours):.0f} ชม. ({float(ot.amount):,.0f} บาท) รอการอนุมัติ')

        from .serializers import OTRequestSerializer as _S
        return Response(_S(ot).data)


# ─── Holidays ────────────────────────────────────────────────────────────────

class HolidayViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = HolidaySerializer

    def get_queryset(self):
        qs = Holiday.objects.all()
        year = self.request.query_params.get('year')
        if year:
            qs = qs.filter(year=year)
        return qs

    def perform_create(self, serializer):
        h = serializer.save(is_system=False)
        log_action(self.request.user, f'เพิ่มวันหยุด {h.name}', 'Holiday', h.id, request=self.request)

    def perform_destroy(self, instance):
        if instance.is_system and self.request.user.role != 'admin':
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('เฉพาะแอดมินเท่านั้นที่สามารถลบวันหยุดราชการได้')
        log_action(self.request.user, f'ลบวันหยุด {instance.name}', 'Holiday', instance.id, request=self.request)
        instance.delete()


# ─── System Settings ─────────────────────────────────────────────────────────

@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def settings_view(request):
    obj, _ = SystemSettings.objects.get_or_create(pk=1)
    if request.method == 'GET':
        return Response(SystemSettingsSerializer(obj).data)
    serializer = SystemSettingsSerializer(obj, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save(updated_by=request.user)
        log_action(request.user, 'อัปเดตตั้งค่าระบบ', request=request)
        return Response(serializer.data)
    return Response(serializer.errors, status=400)


# ─── Admin: fix user departments ──────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def test_tu_api(request):
    """ทดสอบการเชื่อมต่อ TU API ด้วย emp_id ที่ระบุ"""
    if request.user.role != 'admin':
        return Response({'error': 'Permission denied'}, status=403)
    emp_id = request.data.get('emp_id', '0001')
    from .tu_api_service import fetch_employee, clear_cache
    clear_cache()
    result = fetch_employee(str(emp_id))
    if result:
        return Response(result)
    return Response({'error': f'ไม่พบข้อมูลพนักงาน {emp_id} หรือเชื่อมต่อ TU API ไม่ได้'}, status=404)


@api_view(['POST'])
@permission_classes([AllowAny])
def test_tu_auth(request):
    """
    Debug endpoint: ทดสอบ TU Auth API โดยตรง
    POST { "username": "...", "password": "..." }
    คืน raw response จาก TU API (สำหรับ debug เท่านั้น)
    """
    import urllib.request, urllib.error, json as _json
    from .models import SystemSettings

    username = request.data.get('username', '').strip()
    password = request.data.get('password', '').strip()
    if not username or not password:
        return Response({'error': 'กรุณาส่ง username และ password'}, status=400)

    s = SystemSettings.objects.first()
    if not s or not s.tu_api_key:
        return Response({'error': 'ยังไม่ได้ตั้งค่า TU API Key'}, status=400)

    url = 'https://restapi.tu.ac.th/api/v1/auth/Ad/verify'
    body = _json.dumps({'UserName': username, 'PassWord': password}).encode()
    headers = {'Content-Type': 'application/json', 'Application-Key': s.tu_api_key}

    try:
        req = urllib.request.Request(url, data=body, headers=headers, method='POST')
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw = _json.loads(resp.read().decode())
        return Response({'tu_response': raw, 'api_key_prefix': s.tu_api_key[:10] + '...'})
    except urllib.error.HTTPError as e:
        body_text = e.read().decode() if hasattr(e, 'read') else ''
        return Response({'error': f'HTTP {e.code}', 'detail': body_text}, status=500)
    except Exception as e:
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def fix_user_departments(request):
    """Admin assigns all users with department=None to a given department."""
    if request.user.role not in ('admin', 'depthead'):
        return Response({'error': 'Permission denied'}, status=403)
    dept_id = request.data.get('department_id')
    if not dept_id:
        return Response({'error': 'department_id required'}, status=400)
    try:
        from .models import Department
        dept = Department.objects.get(id=int(dept_id))
    except (Department.DoesNotExist, ValueError):
        return Response({'error': 'ไม่พบแผนก'}, status=404)
    updated = User.objects.filter(department__isnull=True, role='staff').update(department=dept)
    return Response({'updated': updated, 'department': dept.name})


# ─── Staff Roster Import ──────────────────────────────────────────────────────

HONORIFICS = [
    'รองศาสตราจารย์ ดร.', 'ผู้ช่วยศาสตราจารย์ ดร.', 'รองศาสตราจารย์',
    'ศาสตราจารย์ ดร.', 'ศาสตราจารย์', 'ว่าที่ ร.ต.', 'ว่าที่ร.ต.',
    'อาจารย์ ดร.', 'อาจารย์', 'นางสาว', 'นาง', 'นาย', 'ดร.',
]

def _strip_honorific(name: str):
    for h in sorted(HONORIFICS, key=len, reverse=True):
        if name.startswith(h):
            name = name[len(h):].strip()
            break
    parts = name.split()
    return (parts[0] if parts else name), (' '.join(parts[1:]) if len(parts) > 1 else '')


def _cell_str(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _thai_first_clusters(name, n=3):
    clusters = []
    for ch in name:
        if clusters and unicodedata.category(ch) in ('Mn', 'Mc'):
            clusters[-1] += ch
        else:
            clusters.append(ch)
    return ''.join(clusters[:n])


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_staff_roster(request):
    """
    Admin อัปโหลด Excel รายชื่อบุคลากร → สร้าง/อัปเดต User accounts

    Format ใหม่ (ตรวจจาก header row):
        employee_id | ชื่อ (ไฟล์รหัส) | สังกัด | ชื่อ (ไฟล์อีเมล) | email_reg | email_tu | หมายเหตุ

    Format เก่า (backward compatible):
        col A=ลำดับ, col B=ชื่อ-สกุล, col C=ตำแหน่ง
        แถว section header = col A ว่าง + col B = ชื่องาน
    """
    if request.user.role != 'admin':
        return Response({'error': 'Permission denied'}, status=403)
    if 'file' not in request.FILES:
        return Response({'error': 'กรุณาแนบไฟล์ Excel'}, status=400)

    default_password = request.data.get('password', 'tustaff2025')
    default_role     = request.data.get('role', 'staff')

    try:
        wb = openpyxl.load_workbook(request.FILES['file'])
    except Exception as e:
        return Response({'error': f'อ่านไฟล์ไม่ได้: {e}'}, status=400)

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return Response({'error': 'ไฟล์ไม่มีข้อมูล'}, status=400)

    # ── ตรวจ format จาก header row ──────────────────────────────────────────
    _NEW_REQUIRED = ['employee_id', 'ชื่อ (ไฟล์รหัส)', 'สังกัด', 'email_reg', 'email_tu']
    header_row = [_cell_str(c) for c in rows[0]]
    is_new_format = all(h in header_row for h in _NEW_REQUIRED)

    created, skipped, errors = [], [], []

    if is_new_format:
        # ── New format: employee_id mapping ──────────────────────────────────
        _ALL_COLS = ['employee_id', 'ชื่อ (ไฟล์รหัส)', 'สังกัด', 'ชื่อ (ไฟล์อีเมล)', 'email_reg', 'email_tu', 'หมายเหตุ']
        col = {h: header_row.index(h) for h in _ALL_COLS if h in header_row}
        data_rows = rows[1:]

        # ขั้น 1: รวบรวมชื่อแผนกทั้งหมด แล้ว get_or_create
        seen_depts = []
        for row in data_rows:
            if not any(row):
                continue
            dn = _cell_str(row[col['สังกัด']]) if 'สังกัด' in col and col['สังกัด'] < len(row) else ''
            if dn and dn not in seen_depts:
                seen_depts.append(dn)

        used_codes = set(Department.objects.values_list('code', flat=True))
        dept_map = {}
        for dn in seen_depts:
            dept = Department.objects.filter(name=dn).first()
            if not dept:
                prefix = _thai_first_clusters(dn, 3)
                n = 1
                while f'{prefix}_{n}' in used_codes:
                    n += 1
                code = f'{prefix}_{n}'
                used_codes.add(code)
                dept = Department.objects.create(name=dn, code=code)
            dept_map[dn] = dept

        # ขั้น 2: สร้าง/อัปเดต User
        updated = []
        for row in data_rows:
            if not any(row):
                continue
            emp_id = _cell_str(row[col['employee_id']]) if 'employee_id' in col and col['employee_id'] < len(row) else ''
            if not emp_id:
                continue

            full_name = _cell_str(row[col['ชื่อ (ไฟล์รหัส)']]) if 'ชื่อ (ไฟล์รหัส)' in col else ''
            dept_name = _cell_str(row[col['สังกัด']])           if 'สังกัด' in col else ''
            email_reg = _cell_str(row[col['email_reg']])        if 'email_reg' in col else ''
            email_tu  = _cell_str(row[col['email_tu']])         if 'email_tu' in col else ''

            first, last = _strip_honorific(full_name)
            dept = dept_map.get(dept_name)

            existing = User.objects.filter(employee_id=emp_id).first()
            if existing:
                existing.username   = emp_id
                existing.first_name = first
                existing.last_name  = last
                existing.email      = email_tu
                existing.department = dept
                if hasattr(existing, 'notify_email'):
                    existing.notify_email = email_reg
                existing.save()
                updated.append({'username': emp_id, 'name': f'{first} {last}', 'department': dept.name if dept else ''})
            else:
                try:
                    u = User(
                        username=emp_id,
                        employee_id=emp_id,
                        first_name=first,
                        last_name=last,
                        email=email_tu,
                        role=default_role,
                        department=dept,
                        is_active=True,
                    )
                    if hasattr(u, 'notify_email'):
                        u.notify_email = email_reg
                    u.set_password(emp_id)
                    u.save()
                    created.append({'username': emp_id, 'name': f'{first} {last}', 'department': dept.name if dept else '', 'email': email_tu})
                except Exception as e:
                    errors.append({'name': f'{first} {last}', 'error': str(e)})

        log_action(request.user, f'import staff roster (new format): สร้าง {len(created)} อัปเดต {len(updated)} บัญชี', request=request)
        return Response({
            'format':        'new',
            'created':       len(created),
            'updated':       len(updated),
            'errors':        len(errors),
            'users':         created,
            'updated_users': updated,
        })

    else:
        # ── Old format: ลำดับ | ชื่อ-สกุล | ตำแหน่ง ─────────────────────────
        current_dept = None

        for row in rows:
            seq, name_raw, position = (row + (None, None, None))[:3]
            if not name_raw:
                continue
            name_raw = str(name_raw).strip()
            if name_raw in ('ชื่อ-สกุล',) or 'รายชื่อบุคลากร' in name_raw:
                continue

            # section header
            if seq is None:
                existing = Department.objects.filter(name=name_raw).first()
                if existing:
                    current_dept = existing
                else:
                    import time as _time
                    base_code = name_raw[:15].upper().replace(' ', '_')
                    code = base_code
                    if Department.objects.filter(code=code).exists():
                        code = f'{base_code[:10]}_{int(_time.time()) % 10000}'
                    current_dept = Department.objects.create(name=name_raw, code=code)
                continue

            # employee row
            first, last = _strip_honorific(name_raw)
            clean = re.sub(r'[^฀-๿a-zA-Z0-9]', '', first)
            username = f'{clean}_{seq}'

            if User.objects.filter(username=username).exists():
                skipped.append({'username': username, 'name': f'{first} {last}'})
                continue

            try:
                u = User(
                    username=username,
                    first_name=first,
                    last_name=last,
                    email='',
                    role=default_role,
                    department=current_dept,
                    is_active=True,
                )
                u.set_password(default_password)
                u.save()
                created.append({
                    'username': username,
                    'name': f'{first} {last}',
                    'department': current_dept.name if current_dept else '',
                    'position': str(position or '').strip(),
                })
            except Exception as e:
                errors.append({'name': name_raw, 'error': str(e)})

        log_action(request.user, f'import staff roster: สร้าง {len(created)} บัญชี', request=request)
        return Response({
            'format':           'legacy',
            'created':          len(created),
            'skipped':          len(skipped),
            'errors':           len(errors),
            'users':            created,
            'default_password': default_password,
        })


# ─── Time Log Import ──────────────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_timelog(request):
    if 'file' not in request.FILES:
        return Response({'error': 'กรุณาแนบไฟล์ Excel'}, status=400)

    f = request.FILES['file']

    # รับ department_id จาก form (ถ้ามี) เพื่อ assign ให้ user ที่สร้างใหม่
    import_dept = None
    dept_id_param = request.data.get('department_id') or request.data.get('department')
    if dept_id_param:
        try:
            from .models import Department
            import_dept = Department.objects.get(id=int(dept_id_param))
        except (Department.DoesNotExist, ValueError, TypeError):
            pass
    # fallback: ใช้แผนกของ admin ที่ import (ถ้ามี)
    if import_dept is None and request.user.department:
        import_dept = request.user.department

    history = ImportHistory.objects.create(
        filename=f.name, imported_by=request.user, status='partial'
    )

    # ── helpers ──────────────────────────────────────────────
    def fmt_time(t):
        if t is None:
            return ''
        if hasattr(t, 'strftime'):
            try:
                return t.strftime('%H:%M')
            except Exception:
                return ''
        s = str(t).strip()
        # If it looks like a datetime string "2026-03-01 08:30:00", extract time
        if len(s) > 10 and ' ' in s:
            s = s.split(' ')[1]
        return s[:5] if len(s) >= 5 else s

    def fmt_date(d):
        if d is None:
            return ''
        if hasattr(d, 'strftime'):
            try:
                return d.strftime('%Y-%m-%d')
            except Exception:
                return str(d)
        return str(d).strip()[:10]

    def calc_ot(in_str, out_str, time_period='', day_type='weekday'):
        try:
            ih, im = map(int, in_str.replace('.', ':').split(':')[:2])
            oh, om = map(int, out_str.replace('.', ':').split(':')[:2])
            in_mins  = ih * 60 + im
            out_mins = oh * 60 + om

            if day_type == 'holiday':
                worked_mins = max(0, out_mins - in_mins)
                # หักพักเที่ยง: ถ้าช่วงเข้า-ออกทับกับ 12:00-13:00
                overlap_start = max(in_mins, 12 * 60)
                overlap_end   = min(out_mins, 13 * 60)
                overlap = max(0, overlap_end - overlap_start)
                worked_mins -= overlap
                ot_hours = min(7, worked_mins // 60)
                return float(ot_hours)
            else:
                ot_start = 16 * 60 if time_period == 'เช้า' else 16 * 60 + 30
                ot_mins = max(0, out_mins - ot_start)
                ot_hours = ot_mins // 60
                return float(ot_hours)
        except Exception:
            return 0.0

    def looks_like_id(val):
        """Return True if val looks like an employee ID (numeric or EMP-xxx format)."""
        if val is None:
            return False
        s = str(val).strip()
        if not s:
            return False
        # pure numeric (0004, 1024, etc.)
        if s.isdigit():
            return True
        # EMP-xxx, AD-xxx, CHK-xxx, etc.
        if re.match(r'^[A-Z]{2,4}-\d+$', s):
            return True
        return False

    # ── auto-detect column layout from header row ─────────────
    try:
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        # Keyword maps for column detection
        ID_KEYS     = {'id', 'รหัส', 'emp id', 'employee id', 'รหัสพนักงาน', 'no', 'no.', 'เลขที่'}
        NAME_KEYS   = {'name', 'ชื่อ', 'ชื่อ-สกุล', 'ชื่อพนักงาน', 'ชื่อ สกุล'}
        DATE_KEYS   = {'date', 'วันที่', 'วัน/เดือน/ปี'}
        IN_KEYS     = {'in', 'check in', 'เข้า', 'เวลาเข้า', 'check-in', 'time in', 'in time'}
        OUT_KEYS    = {'out', 'check out', 'ออก', 'เวลาออก', 'check-out', 'time out', 'out time'}
        STATUS_KEYS      = {'attendance status', 'status', 'สถานะ', 'attendance', 'หมายเหตุ', 'note', 'remark'}
        TIME_PERIOD_KEYS = {'time period', 'กะ', 'กะการทำงาน', 'shift', 'period'}

        col_id = col_name = col_date = col_in = col_out = col_status = col_time_period = None
        data_start_row = 2  # default: start from row 2

        all_rows = list(ws.iter_rows(min_row=1, max_row=15, values_only=True))
        for row_idx, row in enumerate(all_rows):
            clean = [str(v).strip().lower() if v is not None else '' for v in row]
            # Check if this row looks like a header row
            matches = 0
            tmp_id = tmp_name = tmp_date = tmp_in = tmp_out = tmp_status = tmp_time_period = None
            for ci, cell in enumerate(clean):
                if cell in ID_KEYS and tmp_id is None:
                    tmp_id = ci; matches += 1
                elif cell in NAME_KEYS and tmp_name is None:
                    tmp_name = ci; matches += 1
                elif cell in DATE_KEYS and tmp_date is None:
                    tmp_date = ci; matches += 1
                elif cell in IN_KEYS and tmp_in is None:
                    tmp_in = ci; matches += 1
                elif cell in OUT_KEYS and tmp_out is None:
                    tmp_out = ci; matches += 1
                elif cell in STATUS_KEYS and tmp_status is None:
                    tmp_status = ci
                elif cell in TIME_PERIOD_KEYS and tmp_time_period is None:
                    tmp_time_period = ci
            if matches >= 2:
                col_id          = tmp_id
                col_name        = tmp_name
                col_date        = tmp_date
                col_in          = tmp_in
                col_out         = tmp_out
                col_status      = tmp_status
                col_time_period = tmp_time_period
                data_start_row = row_idx + 2  # next row (1-indexed)
                break

        # ถ้า col_time_period ยังไม่พบ (header ว่าง/None) — ลองตรวจจากค่า data
        if col_time_period is None and col_id is not None:
            SHIFT_VALS = {'ปกติ', 'เช้า', 'กลางวัน', 'บ่าย', 'ดึก', 'เย็น'}
            for _row in ws.iter_rows(min_row=data_start_row, max_row=data_start_row + 10, values_only=True):
                for _ci, _cv in enumerate(_row):
                    if _cv and str(_cv).strip() in SHIFT_VALS:
                        col_time_period = _ci
                        break
                if col_time_period is not None:
                    break

        # Fallback: guess by position if auto-detect failed
        # Try to sniff from first data-like row
        if col_id is None:
            for row in ws.iter_rows(min_row=data_start_row, max_row=data_start_row + 5, values_only=True):
                if not any(row):
                    continue
                # Try col0=id or col1=id
                if looks_like_id(row[0]):
                    col_id = 0
                    # guess: id, name, date, weekday, in, out, overtime, status
                    col_name = None; col_date = 2; col_in = 4; col_out = 5; col_status = 7
                elif len(row) > 1 and looks_like_id(row[1]):
                    col_id = 1
                    col_name = 0; col_date = 2; col_in = 4; col_out = 5; col_status = 7
                break
            if col_id is None:
                col_id = 0; col_name = None; col_date = 2; col_in = 4; col_out = 5; col_status = 7

        total = success = errors = skipped_users = 0
        error_lines = []
        rows_out = []
        row_id = 1

        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if not any(row):
                continue

            def gcol(idx):
                return row[idx] if idx is not None and len(row) > idx else None

            emp_id      = gcol(col_id)
            excel_name  = gcol(col_name)
            date_val    = gcol(col_date)
            check_in    = gcol(col_in)
            check_out   = gcol(col_out)
            att_status       = str(gcol(col_status) or '').strip()
            time_period_val  = str(gcol(col_time_period) or '').strip()

            if isinstance(check_in, str) and check_in.strip() in ('-', '--', '', 'None', '-  -'):
                check_in = None
            if isinstance(check_out, str) and check_out.strip() in ('-', '--', '', 'None', '-  -'):
                check_out = None

            # Skip rows where emp_id looks like a header or is empty
            if not looks_like_id(emp_id):
                continue

            date_str  = fmt_date(date_val)
            in_str    = fmt_time(check_in)
            out_str   = fmt_time(check_out)

            import datetime as _dt
            row_day_type = 'weekday'
            row_holiday_name = ''
            row_holiday_type = ''
            _d = None
            try:
                _d = _dt.date.fromisoformat(date_str)
                _h_obj = Holiday.objects.filter(date=_d).first()
                _is_wknd = _d.weekday() >= 5
                if _h_obj or _is_wknd:
                    row_day_type = 'holiday'
                    if _h_obj:
                        row_holiday_name = _h_obj.name
                        row_holiday_type = _h_obj.holiday_type
                    else:
                        row_holiday_name = 'เสาร์-อาทิตย์'
                        row_holiday_type = 'weekend'
            except (ValueError, TypeError):
                pass

            # วันเสาร์-อาทิตย์ที่ไม่มีการเข้า-ออกงาน ถือว่าปกติ ไม่ใช่ "ขาด"
            if _d is not None and _d.weekday() >= 5 and not check_in and not check_out:
                att_status = ''

            ot_val = calc_ot(in_str, out_str, time_period_val, row_day_type) if (in_str and out_str) else 0.0
            flag   = (row_day_type == 'weekday' and (not in_str or not out_str)) or ot_val > 8

            total += 1

            # ดึงข้อมูลจาก TU API (ถ้าเปิดใช้งาน) หรือใช้จาก Excel
            from .tu_api_service import fetch_employee
            tu_data = fetch_employee(str(emp_id))  # คืน dict หรือ None

            if tu_data:
                # ใช้ข้อมูลจาก TU API (น่าเชื่อถือกว่า)
                first = tu_data['first_name']
                last  = tu_data['last_name']
                # dept จาก TU API มีความสำคัญกว่า import_dept ที่ admin เลือก
                dept_name = tu_data['dept_name']
                resolved_dept = Department.objects.get_or_create(
                    name=dept_name, defaults={'code': dept_name[:15].upper().replace(' ', '_')}
                )[0] if dept_name else import_dept
            else:
                # fallback: ใช้ชื่อจาก Excel
                raw_name = str(excel_name).strip() if excel_name else ''
                parts = raw_name.split() if raw_name else []
                first = parts[0] if parts else str(emp_id)
                last  = ' '.join(parts[1:]) if len(parts) > 1 else ''
                resolved_dept = import_dept

            # Try to match user in DB — skip row if not found
            try:
                user = User.objects.get(employee_id=str(emp_id))
                # อัปเดตข้อมูลที่อาจเปลี่ยนแปลง (ชื่อ/แผนก)
                changed = False
                if tu_data and user.first_name != first:
                    user.first_name = first; changed = True
                if tu_data and user.last_name != last:
                    user.last_name = last;  changed = True
                if user.department is None and resolved_dept:
                    user.department = resolved_dept; changed = True
                if tu_data and tu_data.get('email') and not user.email:
                    user.email = tu_data['email']; changed = True
                if changed:
                    user.save()
            except User.DoesNotExist:
                skipped_users += 1
                continue

            success += 1
            db_name   = (f'{user.first_name} {user.last_name}'.strip()) or user.username
            dept_name = (user.department.name if getattr(user, 'department', None) else None) or 'ไม่ระบุ'
            try:
                TimeLog.objects.update_or_create(
                    user=user, log_date=date_val or date_str,
                    defaults={'check_in': check_in, 'check_out': check_out, 'source_file': f.name,
                              'time_period': time_period_val, 'attendance_status': att_status}
                )
            except Exception as e:
                error_lines.append(f'แถว {total}: บันทึก timelog ไม่ได้ — {e}')

            rows_out.append({
                'id': row_id, 'date': date_str,
                'empId': str(emp_id), 'name': db_name,
                'dept': dept_name, 'in': in_str, 'out': out_str,
                'ot': str(ot_val), 'flag': flag,
                'timePeriod':   time_period_val,
                'attendanceStatus': att_status,
                'dayType':      row_day_type,
                'holidayName':  row_holiday_name,
                'holidayType':  row_holiday_type,
            })
            row_id += 1

        history.total_rows = total
        history.success_rows = success
        history.error_rows = errors
        history.error_detail = '\n'.join(error_lines[:50])
        history.status = 'success' if errors == 0 else ('failed' if success == 0 else 'partial')
        history.rows_data = rows_out  # persist all rows regardless of user match
        history.save()

        from .tu_api_service import _CACHE
        _CACHE.clear()

        log_action(request.user, f'นำเข้าไฟล์ {f.name} ({success}/{total} รายการ, ข้าม {skipped_users} รหัสไม่พบ)', request=request)
        resp = dict(ImportHistorySerializer(history).data)
        resp['rows'] = rows_out
        resp['skipped_users'] = skipped_users
        return Response(resp)

    except Exception as e:
        from .tu_api_service import _CACHE
        _CACHE.clear()
        history.status = 'failed'
        history.error_detail = str(e)
        history.save()
        return Response({'error': str(e)}, status=400)


# ─── Import History & Audit Log ───────────────────────────────────────────────

class ImportHistoryViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = ImportHistorySerializer
    queryset = ImportHistory.objects.all().order_by('-imported_at')


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = AuditLogSerializer
    queryset = AuditLog.objects.filter(user__role='admin').order_by('-created_at')


# ─── Timelog & Dashboard API ──────────────────────────────────────────────────

def _thai_to_greg(month_str):
    """Convert "2569-03" (Thai year) to (2026, 3)."""
    try:
        thai_year, mon = month_str.split('-')
        return int(thai_year) - 543, int(mon)
    except Exception:
        return None, None


def _calc_ot(check_in, check_out, time_period='', day_type='weekday'):
    """Return OT hours (float) from time fields."""
    try:
        in_mins  = check_in.hour * 60 + check_in.minute
        out_mins = check_out.hour * 60 + check_out.minute
        if day_type == 'holiday':
            worked_mins = max(0, out_mins - in_mins)
            overlap_start = max(in_mins, 12 * 60)
            overlap_end   = min(out_mins, 13 * 60)
            overlap = max(0, overlap_end - overlap_start)
            worked_mins -= overlap
            return float(min(7, worked_mins // 60))
        else:
            ot_start = 16 * 60 if time_period == 'เช้า' else 16 * 60 + 30
            ot_mins = max(0, out_mins - ot_start)
            return float(min(4, ot_mins // 60))
    except Exception:
        return 0.0


def _row_from_timelog(idx, tl):
    in_str   = tl.check_in.strftime('%H:%M')  if tl.check_in  else ''
    out_str  = tl.check_out.strftime('%H:%M') if tl.check_out else ''
    h_obj    = Holiday.objects.filter(date=tl.log_date).first()
    is_weekend = tl.log_date.weekday() >= 5
    day_type = 'holiday' if (h_obj or is_weekend) else 'weekday'
    ot_val   = _calc_ot(tl.check_in, tl.check_out, tl.time_period, day_type) if (tl.check_in and tl.check_out) else 0.0
    flag     = (day_type == 'weekday' and (not in_str or not out_str)) or ot_val > 8
    return {
        'id':    idx,
        'date':  str(tl.log_date),
        'empId': tl.user.employee_id or '',
        'name':  tl.user.get_full_name() or tl.user.username,
        'dept':  tl.user.department.name if tl.user.department else 'ไม่ระบุ',
        'in':    in_str,
        'out':   out_str,
        'ot':               str(ot_val),
        'flag':             flag,
        'attendanceStatus': tl.attendance_status or '',
        'timePeriod':       tl.time_period or '',
        'dayType':          day_type,
        'holidayName':      h_obj.name if h_obj else ('เสาร์-อาทิตย์' if is_weekend else ''),
        'holidayType':      h_obj.holiday_type if h_obj else ('weekend' if is_weekend else ''),
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def timelog_list_view(request):
    """Admin: list all timelogs for a given Thai-calendar month.
    Falls back to ImportHistory.rows_data if no TimeLog rows saved (unmatched users).
    """
    month_param = request.query_params.get('month', '')
    greg_year, mon = _thai_to_greg(month_param)

    qs = TimeLog.objects.select_related('user', 'user__department')
    if greg_year and mon:
        qs = qs.filter(log_date__year=greg_year, log_date__month=mon)
    qs = qs.order_by('user__employee_id', 'log_date')

    rows = [_row_from_timelog(i + 1, tl) for i, tl in enumerate(qs)]

    # Fallback: if no TimeLog rows but ImportHistory has raw data for this month
    if not rows and greg_year and mon:
        from datetime import date as _date
        import calendar
        last_day = calendar.monthrange(greg_year, mon)[1]
        month_start = _date(greg_year, mon, 1)
        month_end   = _date(greg_year, mon, last_day)
        hist = (ImportHistory.objects
                .filter(rows_data__isnull=False, imported_at__date__gte=month_start - __import__('datetime').timedelta(days=400))
                .order_by('-imported_at')
                .first())
        # Find the import whose first row date matches the month
        if hist and hist.rows_data:
            sample_date = (hist.rows_data[0].get('date', '') or '')[:7]  # 'YYYY-MM'
            if sample_date == f'{greg_year}-{mon:02d}':
                rows = hist.rows_data

    rows = _enrich_rows_day_type(rows)
    return Response({'rows': rows, 'total': len(rows)})


def _enrich_rows_day_type(rows):
    """Recompute dayType/holidayName/holidayType for each row from current Holiday DB.
    Used to fix stale rows_data stored in ImportHistory before this field existed.
    """
    import datetime as _dt
    enriched = []
    for r in rows:
        date_str = r.get('date', '')
        try:
            _d = _dt.date.fromisoformat(date_str)
            h_obj = Holiday.objects.filter(date=_d).first()
            is_wknd = _d.weekday() >= 5
            if h_obj or is_wknd:
                day_type = 'holiday'
                h_name = h_obj.name if h_obj else 'เสาร์-อาทิตย์'
                h_type = h_obj.holiday_type if h_obj else 'weekend'
            else:
                day_type = 'weekday'
                h_name = ''
                h_type = ''
        except (ValueError, TypeError):
            day_type = r.get('dayType', 'weekday')
            h_name = r.get('holidayName', '')
            h_type = r.get('holidayType', '')
        enriched.append({**r, 'dayType': day_type, 'holidayName': h_name, 'holidayType': h_type})
    return enriched


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def timelog_my_view(request):
    """Staff: list own timelogs for a given Thai-calendar month.
    Falls back to ImportHistory.rows_data filtered by employee_id."""
    month_param = request.query_params.get('month', '')
    greg_year, mon = _thai_to_greg(month_param)

    qs = TimeLog.objects.filter(user=request.user).select_related('user', 'user__department')
    if greg_year and mon:
        qs = qs.filter(log_date__year=greg_year, log_date__month=mon)
    qs = qs.order_by('log_date')

    rows = [_row_from_timelog(i + 1, tl) for i, tl in enumerate(qs)]

    # Fallback: use ImportHistory rows filtered by employee_id
    if not rows and greg_year and mon:
        emp_id = request.user.employee_id or request.user.username
        hist = (ImportHistory.objects
                .filter(rows_data__isnull=False)
                .order_by('-imported_at').first())
        if hist and hist.rows_data:
            month_prefix = f'{greg_year}-{mon:02d}'
            my_rows = [r for r in hist.rows_data
                       if str(r.get('empId', '')) == str(emp_id)
                       and str(r.get('date', '')).startswith(month_prefix)]
            if my_rows:
                rows = my_rows

    rows = _enrich_rows_day_type(rows)

    total_ot = sum(float(r.get('ot', 0)) for r in rows)
    flag_count = sum(1 for r in rows if r.get('flag'))
    total_ot_baht = sum(
        float(r.get('ot', 0)) * (70 if r.get('dayType') == 'holiday' else 60)
        for r in rows
    )
    return Response({
        'rows': rows,
        'total': len(rows),
        'total_ot': round(total_ot, 2),
        'total_ot_baht': round(total_ot_baht, 2),
        'flag_count': flag_count,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def seed_holidays_view(request):
    """รับ holiday data จาก frontend แล้ว upsert ลง DB
    Body: { year: 2569, holidays: [{date, name, holiday_type?, is_system?}, ...] }
    """
    import datetime as _dt

    thai_year = request.data.get('year')
    holidays_data = request.data.get('holidays', [])

    if not thai_year:
        return Response({'error': 'กรุณาระบุปี พ.ศ.'}, status=400)
    if not isinstance(holidays_data, list) or len(holidays_data) == 0:
        return Response({'error': 'ไม่มีข้อมูลวันหยุด'}, status=400)
    try:
        thai_year = int(thai_year)
    except (ValueError, TypeError):
        return Response({'error': 'ปีไม่ถูกต้อง'}, status=400)

    created_count = 0
    updated_count = 0
    for h in holidays_data:
        try:
            date_obj = _dt.date.fromisoformat(h['date'])
        except (ValueError, KeyError):
            continue
        name = (h.get('localName') or h.get('name', '')).strip()
        holiday_type = h.get('holiday_type', 'official')
        obj, was_created = Holiday.objects.get_or_create(
            date=date_obj,
            defaults={
                'name':         name,
                'holiday_type': holiday_type,
                'year':         thai_year,
                'is_system':    True,
            }
        )
        if was_created:
            created_count += 1
        elif obj.is_system and obj.name != name:
            obj.name = name
            obj.save(update_fields=['name'])
            updated_count += 1

    log_action(request.user, f'sync วันหยุดราชการ พ.ศ. {thai_year} ({created_count} ใหม่)', 'Holiday', None, request=request)
    return Response({
        'year':    thai_year,
        'created': created_count,
        'updated': updated_count,
        'message': f'ซิงค์วันหยุด {len(holidays_data)} รายการ (ใหม่ {created_count}, อัปเดต {updated_count})',
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_summary_view(request):
    """Admin dashboard KPIs."""
    from datetime import date as _date
    today = _date.today()

    total_users   = User.objects.filter(is_active=True).exclude(role='admin').count()
    latest_import = ImportHistory.objects.order_by('-imported_at').first()
    ot_count      = OTRequest.objects.filter(
        created_at__year=today.year, created_at__month=today.month
    ).count()

    # Available imported months — from TimeLog first, then fall back to ImportHistory.rows_data
    from django.db.models.functions import TruncMonth
    months = (TimeLog.objects
              .annotate(ym=TruncMonth('log_date'))
              .values_list('ym', flat=True)
              .distinct()
              .order_by('-ym'))
    thai_months = []
    for m in months:
        if m:
            thai_months.append(f'{m.year + 543}-{m.month:02d}')

    # If no TimeLogs, derive months from ImportHistory raw data
    if not thai_months:
        for hist in ImportHistory.objects.filter(rows_data__isnull=False).order_by('-imported_at'):
            if hist.rows_data:
                sample_date = (hist.rows_data[0].get('date', '') or '')
                if len(sample_date) >= 7:
                    y_str, m_str = sample_date[:4], sample_date[5:7]
                    tag = f'{int(y_str) + 543}-{m_str}'
                    if tag not in thai_months:
                        thai_months.append(tag)

    # Department user distribution
    from django.db.models import Count as _Count
    dept_dist = list(
        User.objects
        .filter(is_active=True, department__isnull=False)
        .exclude(role='admin')
        .values('department__id', 'department__name')
        .annotate(count=_Count('id'))
        .order_by('-count')
    )

    return Response({
        'total_users':    total_users,
        'ot_requests':    ot_count,
        'system_status':  'ปกติ',
        'imported_months': thai_months,
        'dept_distribution': [
            {'id': d['department__id'], 'name': d['department__name'], 'count': d['count']}
            for d in dept_dist
        ],
        'latest_import': {
            'date':     latest_import.imported_at.strftime('%d/%m/%Y %H:%M') if latest_import else None,
            'by':       latest_import.imported_by.get_full_name()            if latest_import else None,
            'filename': latest_import.filename                               if latest_import else None,
            'total':    latest_import.total_rows                             if latest_import else 0,
        } if latest_import else None,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def staff_summary_view(request):
    """Staff dashboard KPIs: their OT totals & request statuses."""
    month_param = request.query_params.get('month', '')
    greg_year, mon = _thai_to_greg(month_param)

    tl_qs = TimeLog.objects.filter(user=request.user)
    if greg_year and mon:
        tl_qs = tl_qs.filter(log_date__year=greg_year, log_date__month=mon)

    total_ot = 0.0
    for tl in tl_qs:
        day_type = 'holiday' if (Holiday.objects.filter(date=tl.log_date).exists() or tl.log_date.weekday() >= 5) else 'weekday'
        total_ot += _calc_ot(tl.check_in, tl.check_out, tl.time_period, day_type) if (tl.check_in and tl.check_out) else 0.0

    # Fallback: use ImportHistory rows_data if no TimeLog entries
    if tl_qs.count() == 0 and greg_year and mon:
        emp_id = request.user.employee_id or request.user.username
        hist = ImportHistory.objects.filter(rows_data__isnull=False).order_by('-imported_at').first()
        if hist and hist.rows_data:
            month_prefix = f'{greg_year}-{mon:02d}'
            my_rows = [r for r in hist.rows_data
                       if str(r.get('empId', '')) == str(emp_id)
                       and str(r.get('date', '')).startswith(month_prefix)]
            total_ot = sum(float(r.get('ot', 0)) for r in my_rows)

    ot_qs = OTRequest.objects.filter(staff=request.user)
    if greg_year and mon:
        ot_qs = ot_qs.filter(work_date__year=greg_year, work_date__month=mon)

    status_counts = {}
    for r in ot_qs.values('status'):
        status_counts[r['status']] = status_counts.get(r['status'], 0) + 1

    total_ot_baht = sum(float(r.amount) for r in ot_qs)
    rejected = ['head_rejected', 'checker_rejected']
    approved = ['checker_approved', 'completed']
    pending  = ['submitted', 'head_approved', 'rep_forwarded']

    return Response({
        'total_ot_hours': round(total_ot, 2),
        'total_ot_baht':  round(total_ot_baht, 2),
        'timelog_days':   tl_qs.count(),
        'ot_total':       ot_qs.count(),
        'ot_submitted':   ot_qs.filter(status__in=pending).count(),
        'ot_approved':    ot_qs.filter(status__in=approved).count(),
        'ot_rejected':    ot_qs.filter(status__in=rejected).count(),
        'ot_draft':       status_counts.get('draft', 0),
    })


# ─── Budget Status (real-time) ───────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def budget_status_view(request):
    """GET /api/budget-status/?department=<id>&month=YYYY-MM
    คืน: budget, used, remaining สำหรับแสดง real-time บน depthead pending
    used = ยอด OT ที่ head_approved ขึ้นไปแล้วในเดือนนั้น
    """
    dept_id  = request.query_params.get('department')
    month_p  = request.query_params.get('month')   # "2026-03"
    if not dept_id or not month_p:
        return Response({'error': 'ต้องระบุ department และ month'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        year, mon = int(month_p[:4]), int(month_p[5:7])
    except (ValueError, IndexError):
        return Response({'error': 'รูปแบบ month ไม่ถูกต้อง (YYYY-MM)'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        mb = DepartmentMonthlyBudget.objects.get(department_id=dept_id, year=year, month=mon)
        budget = float(mb.budget)
    except DepartmentMonthlyBudget.DoesNotExist:
        budget = 0

    used = sum(
        _calc_ot_amount(r.ot_hours, r.day_type)
        for r in OTRequest.objects.filter(
            department_id=dept_id,
            work_date__year=year, work_date__month=mon,
            status__in=['head_approved', 'rep_forwarded', 'checker_approved', 'completed'],
        )
    )
    return Response({
        'budget':    budget,
        'used':      round(used, 2),
        'remaining': round(budget - used, 2) if budget > 0 else None,
    })


# ─── Bulk Approve (checker) ──────────────────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_approve_view(request):
    """POST /api/ot-requests/bulk-approve/
    checker อนุมัติหลาย OT พร้อมกัน + ส่ง notification รวมต่อแผนก (ไม่ spam ทีละรายการ)
    body: { ids: [1,2,...], note: "" }
    """
    effective_role = get_effective_role(request.user, request)
    if effective_role != 'checker':
        return Response({'error': 'สิทธิ์ไม่เพียงพอ'}, status=status.HTTP_403_FORBIDDEN)

    ids  = request.data.get('ids', [])
    note = request.data.get('note', '')
    if not ids:
        return Response({'error': 'ไม่มีรายการที่เลือก'}, status=status.HTTP_400_BAD_REQUEST)

    checker_name  = request.user.get_full_name() or request.user.username
    approved_list = []
    budget_warnings = []

    for ot_id in ids:
        try:
            ot = OTRequest.objects.get(id=ot_id, status='rep_forwarded')
        except OTRequest.DoesNotExist:
            continue

        # ตรวจเพดานงบ
        work_dt = ot.work_date
        year, mon = work_dt.year, work_dt.month
        budget_warning = None
        try:
            mb = DepartmentMonthlyBudget.objects.get(department=ot.staff.department, year=year, month=mon)
            monthly_budget = float(mb.budget)
            if monthly_budget > 0:
                already_used = sum(_calc_ot_amount(r.ot_hours, r.day_type)
                    for r in OTRequest.objects.filter(staff__department=ot.staff.department,
                        work_date__year=year, work_date__month=mon,
                        status__in=['checker_approved', 'completed']))
                new_total = already_used + _calc_ot_amount(ot.ot_hours, ot.day_type)
                if new_total > monthly_budget:
                    over_by = new_total - monthly_budget
                    pct = round(new_total / monthly_budget * 100)
                    budget_warning = (f'⚠ ยอด OT แผนก {ot.staff.department.name} จะเกินเพดานงบ '
                        f'{monthly_budget:,.0f} บาท ({pct}% — เกิน {over_by:,.0f} บาท)')
        except DepartmentMonthlyBudget.DoesNotExist:
            pass

        ot.status              = 'checker_approved'
        ot.checker_approved_by = request.user
        ot.checker_approved_at = timezone.now()
        ot.checker_note        = note
        ot.save()
        log_action(request.user, f'ผู้ตรวจสอบอนุมัติคำร้อง OT #{ot.id}', 'OTRequest', ot.id, request=request)
        approved_list.append(ot)
        if budget_warning:
            budget_warnings.append(budget_warning)

    # ── ส่ง notification รวมต่อแผนก (1 แจ้งเตือนต่อแผนก ไม่ใช่ต่อรายการ) ──
    from collections import defaultdict
    by_dept = defaultdict(list)
    for ot in approved_list:
        by_dept[ot.department_id].append(ot)

    for dept_id, dept_ots in by_dept.items():
        dept      = dept_ots[0].department
        dept_name = dept.name
        count     = len(dept_ots)
        total_amt = sum(float(o.amount) for o in dept_ots)
        month_str = dept_ots[0].work_date.strftime('%-m') if hasattr(dept_ots[0].work_date, 'strftime') else ''
        try:
            import datetime as _dt
            wd = dept_ots[0].work_date
            thai_month = ['ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.'][wd.month - 1]
            thai_year  = wd.year + 543
            period     = f'{thai_month} {thai_year}'
        except Exception:
            period = ''

        msg_depthead = (f'ผู้ตรวจสอบ ({checker_name}) อนุมัติ OT แผนก {dept_name} '
                        f'ประจำ{period} จำนวน {count} รายการ รวม {total_amt:,.0f} บาท')
        msg_deptrep  = msg_depthead

        deptheads = list(User.objects.filter(role='depthead', department_id=dept_id, is_active=True))
        deptreps  = list(User.objects.filter(role='deptrep',  department_id=dept_id, is_active=True))

        # แจ้งพนักงานแต่ละคน (ยังคงรายบุคคลเพราะเป็นเรื่องเงินของตัวเอง)
        staff_map = defaultdict(list)
        for ot in dept_ots:
            staff_map[ot.staff_id].append(ot)
        for staff_id, staff_ots in staff_map.items():
            staff = staff_ots[0].staff
            s_count = len(staff_ots)
            s_amt   = sum(float(o.amount) for o in staff_ots)
            _notify_ot(staff_ots[0], 'ot_checker_approved', [staff],
                f'ผู้ตรวจสอบ ({checker_name}) อนุมัติ OT ของคุณ {s_count} รายการ รวม {s_amt:,.0f} บาท ({period})')

        if deptheads:
            _notify_ot(dept_ots[0], 'ot_checker_approved', deptheads, msg_depthead)
        if deptreps:
            _notify_ot(dept_ots[0], 'ot_checker_approved', deptreps, msg_deptrep)

    resp = {'approved': len(approved_list)}
    if budget_warnings:
        resp['budget_warning'] = budget_warnings[-1]
    return Response(resp)


# ─── Bulk Forward (deptrep → checker) ───────────────────────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def bulk_forward_view(request):
    """POST /api/ot-requests/bulk-forward/
    deptrep ส่งต่อหลายคำร้องพร้อมกัน + ส่งอีเมลสรุปให้ checker ทุกคน
    body: { ids: [1,2,3], note: "..." }
    response: { forwarded: N, notified_emails: [...] }
    """
    effective_role = get_effective_role(request.user, request)
    if effective_role != 'deptrep':
        return Response({'error': 'สิทธิ์ไม่เพียงพอ'}, status=status.HTTP_403_FORBIDDEN)

    # รองรับทั้ง JSON body และ FormData (กรณีส่งพร้อมไฟล์)
    raw_ids = request.data.get('ids', [])
    if isinstance(raw_ids, str):
        import json as _json
        try:
            raw_ids = _json.loads(raw_ids)
        except Exception:
            raw_ids = []
    ids  = raw_ids
    note = request.data.get('note', '')
    doc  = request.FILES.get('document')

    if not ids:
        return Response({'error': 'ไม่มีรายการที่เลือก'}, status=status.HTTP_400_BAD_REQUEST)

    doc = request.FILES.get('document')

    forwarded_list = []
    for ot_id in ids:
        try:
            ot = OTRequest.objects.get(id=ot_id, status='head_approved')
            ot.status           = 'rep_forwarded'
            ot.rep_forwarded_by = request.user
            ot.rep_forwarded_at = timezone.now()
            ot.rep_note         = note
            if doc:
                ot.rep_document = doc
            ot.save()
            log_action(request.user, f'ส่งต่อคำร้อง OT #{ot.id}',
                       'OTRequest', ot.id, request=request)
            forwarded_list.append(ot)
        except OTRequest.DoesNotExist:
            pass

    # แจ้ง checker ทุกคน (in-app + อีเมล)
    notified_emails = []
    if forwarded_list:
        checkers_qs = list(User.objects.filter(role='checker', is_active=True))
        notified_emails = [u.notify_email or u.email for u in checkers_qs if (u.notify_email or u.email)]
        _send_checker_notification(forwarded_list, note, request.user)

        rep_name = request.user.get_full_name() or request.user.username
        dept_name = forwarded_list[0].department.name if forwarded_list else ''
        has_doc = bool(doc)
        doc_text = ' (พร้อมไฟล์เอกสาร)' if has_doc else ''
        # ใช้ OT แรกเป็นตัวแทน (checker จะเห็น batch ทั้งหมดในหน้า dashboard)
        _notify_ot(
            forwarded_list[0], 'ot_rep_forwarded', checkers_qs,
            f'{rep_name} ส่งเอกสาร OT แผนก {dept_name} จำนวน {len(forwarded_list)} รายการ{doc_text} รอการตรวจสอบ',
        )

    return Response({'forwarded': len(forwarded_list), 'notified_emails': notified_emails})


# ─── OT Deadline views ────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def deadline_list_view(request):
    """GET /api/ot-deadline/
    ดึงรายการ deadline ทั้งหมด (admin) หรือ deadline ของเดือนปัจจุบัน (staff)
    query params:
      ?month=2569-06   → ดึงเฉพาะเดือนนั้น
      ?all=1           → ดึงทั้งหมด (admin only)
    """
    import datetime as _dt
    month_param = request.query_params.get('month', '')
    get_all = request.query_params.get('all', '') == '1'

    qs = OTDeadline.objects.all()
    if month_param:
        qs = qs.filter(thai_month=month_param)
    elif not get_all:
        # Default: return only current + next 2 months
        today = _dt.date.today()
        thai_year = today.year + 543
        months = []
        for delta in range(3):
            m = today.month + delta
            y = thai_year
            if m > 12:
                m -= 12
                y += 1
            months.append(f'{y}-{m:02d}')
        qs = qs.filter(thai_month__in=months)

    data = [
        {
            'id':            d.id,
            'thai_month':    d.thai_month,
            'deadline_date': d.deadline_date.isoformat(),
            'note':          d.note,
            'is_passed':     d.deadline_date < _dt.date.today(),
        }
        for d in qs
    ]
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def deadline_upsert_view(request):
    """POST /api/ot-deadline/
    สร้างหรืออัปเดต deadline สำหรับเดือนที่ระบุ (admin only)
    Body: { thai_month: '2569-06', deadline_date: '2026-06-10', note?: '...' }
    """
    import datetime as _dt
    if request.user.role != 'admin':
        return Response({'error': 'ไม่มีสิทธิ์'}, status=403)

    thai_month    = request.data.get('thai_month', '').strip()
    deadline_date = request.data.get('deadline_date', '').strip()
    note          = request.data.get('note', '').strip()

    if not thai_month or not deadline_date:
        return Response({'error': 'กรุณาระบุ thai_month และ deadline_date'}, status=400)

    import re as _re
    if not _re.match(r'^\d{4}-\d{2}$', thai_month):
        return Response({'error': 'รูปแบบ thai_month ต้องเป็น YYYY-MM เช่น 2569-06'}, status=400)

    try:
        date_obj = _dt.date.fromisoformat(deadline_date)
    except ValueError:
        return Response({'error': 'รูปแบบ deadline_date ไม่ถูกต้อง (YYYY-MM-DD)'}, status=400)

    obj, created = OTDeadline.objects.update_or_create(
        thai_month=thai_month,
        defaults={
            'deadline_date': date_obj,
            'note':          note,
            'created_by':    request.user,
        }
    )
    log_action(request.user, 'SET_DEADLINE', 'OTDeadline', str(obj.id),
               f'{thai_month} → {deadline_date}', request)
    return Response({
        'id':            obj.id,
        'thai_month':    obj.thai_month,
        'deadline_date': obj.deadline_date.isoformat(),
        'note':          obj.note,
        'created':       created,
        'message':       f'{"ตั้ง" if created else "อัปเดต"}วันปิดรับ {thai_month} → {deadline_date} เรียบร้อย',
    })


@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def deadline_delete_view(request, pk):
    """DELETE /api/ot-deadline/<pk>/"""
    if request.user.role != 'admin':
        return Response({'error': 'ไม่มีสิทธิ์'}, status=403)
    try:
        obj = OTDeadline.objects.get(pk=pk)
    except OTDeadline.DoesNotExist:
        return Response({'error': 'ไม่พบรายการ'}, status=404)
    thai_month = obj.thai_month
    obj.delete()
    log_action(request.user, 'DELETE_DEADLINE', 'OTDeadline', str(pk), thai_month, request)
    return Response({'message': f'ลบกำหนดวันปิดรับ {thai_month} เรียบร้อย'})


def _check_ot_deadline(thai_month: str) -> 'str | None':
    """ตรวจว่าเดือน thai_month เลย deadline หรือยัง
    Return error message ถ้าเลย, None ถ้าไม่เลยหรือยังไม่ได้ตั้ง deadline
    """
    import datetime as _dt
    try:
        d = OTDeadline.objects.get(thai_month=thai_month)
        if _dt.date.today() > d.deadline_date:
            return f'เลยกำหนดยื่นคำร้องโอทีเดือน {thai_month} แล้ว (ปิดรับวันที่ {d.deadline_date.strftime("%d/%m/%Y")})'
    except OTDeadline.DoesNotExist:
        pass
    return None



# ─── Notification API ──────────────────────────────────────────────────────────

NOTIF_TYPES_BY_ROLE = {
    'staff':     ['ot_head_approved', 'ot_head_rejected', 'ot_rep_forwarded', 'ot_checker_approved', 'ot_checker_rejected'],
    'depthead':  ['ot_submitted', 'ot_checker_approved', 'ot_checker_rejected', 'budget_set', 'no_ot_declared'],
    'deptrep':   ['ot_rep_action_needed', 'ot_checker_approved', 'ot_checker_rejected'],
    'checker':   ['ot_rep_forwarded'],
    'executive': None,  # ทุกประเภท
    'admin':     None,
}

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def notification_list_view(request):
    acting_role = get_effective_role(request.user, request)
    allowed_types = NOTIF_TYPES_BY_ROLE.get(acting_role)
    qs = Notification.objects.filter(recipient=request.user)
    if allowed_types is not None:
        qs = qs.filter(notif_type__in=allowed_types)
    notifs = qs.order_by('-created_at')[:50]
    data = [{
        'id': n.id,
        'message': n.message,
        'notif_type': n.notif_type,
        'ot_request': n.ot_request_id,
        'ot_request_date': n.ot_request.work_date.isoformat() if n.ot_request else None,
        'is_read': n.is_read,
        'created_at': n.created_at.isoformat(),
    } for n in notifs]
    return Response(data)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def notification_mark_read_view(request):
    ids = request.data.get('ids', [])
    qs = Notification.objects.filter(recipient=request.user)
    if ids:
        qs = qs.filter(id__in=ids)
    qs.update(is_read=True)
    return Response({'status': 'ok'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def notification_mark_all_read_view(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return Response({'status': 'ok'})


# ─── DeptHead → DeptRep "ready to forward" notification ───────────────────────

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def notify_rep_ready_view(request):
    """POST /api/notify-rep-ready/  body: {month: "YYYY-MM"}
    หัวหน้ากดปุ่ม "แจ้งตัวแทนว่าพร้อมส่งออก" — ส่ง notification ให้ deptrep ของแผนก
    """
    user = request.user
    effective_role = get_effective_role(user, request)
    if effective_role != 'depthead':
        return Response({'error': 'depthead only'}, status=403)

    month = request.data.get('month', '')  # "YYYY-MM"
    dept = getattr(user, 'department', None)
    if not dept:
        return Response({'error': 'no department'}, status=400)

    deptreps = User.objects.filter(department=dept, role='deptrep', is_active=True)
    if not deptreps.exists():
        deptreps = User.objects.filter(department=dept, extra_roles__contains=['deptrep'], is_active=True)

    if not deptreps.exists():
        return Response({'error': 'ไม่พบตัวแทนแผนกในแผนกนี้'}, status=404)

    # หาตัวอย่าง OTRequest สำหรับ link ใน notification (ใช้รายการแรกของเดือนนั้น)
    ot_ref = None
    if month:
        ot_ref = OTRequest.objects.filter(
            department=dept, status='head_approved',
            work_date__startswith=month
        ).first()

    dept_name = dept.name if dept else ''
    month_label = month[5:7].lstrip('0') if len(month) >= 7 else ''
    year_label = str(int(month[:4]) + 543) if len(month) >= 4 else ''
    msg = f'หัวหน้างาน {dept_name} แจ้งว่าคำร้อง OT เดือน {month_label}/{year_label} พร้อมส่งออกแล้ว'

    for rep in deptreps:
        notif = Notification.objects.create(
            recipient=rep,
            message=msg,
            notif_type='ot_rep_action_needed',
            ot_request=ot_ref,
            is_read=False,
        )
        _push_ws(rep.id, {
            'id': notif.id,
            'message': notif.message,
            'notif_type': notif.notif_type,
            'ot_request': ot_ref.id if ot_ref else None,
            'ot_request_date': str(ot_ref.work_date) if ot_ref else None,
            'is_read': False,
            'created_at': notif.created_at.isoformat(),
        })
        _send_email(rep, f'[SMART OT] {msg[:60]}', msg)

    return Response({'status': 'ok', 'sent_to': deptreps.count()})


# ─── Checker Budget API ────────────────────────────────────────────────────────

THAI_MONTHS_SHORT = ['ม.ค.','ก.พ.','มี.ค.','เม.ย.','พ.ค.','มิ.ย.','ก.ค.','ส.ค.','ก.ย.','ต.ค.','พ.ย.','ธ.ค.']
APPROVED_STATUSES = ['checker_approved', 'completed']


def _calc_ot_amount(ot_hours, day_type):
    return int(float(ot_hours or 0)) * (70 if day_type == 'holiday' else 60)


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def checker_budget_view(request):
    import datetime

    if request.method == 'GET':
        month_param = request.query_params.get('month')
        if not month_param:
            now = datetime.date.today()
            month_param = now.strftime('%Y-%m')
        year, mon = map(int, month_param.split('-'))

        managed_ids = set(
            User.objects.filter(role__in=['depthead', 'deptrep'], is_active=True)
            .values_list('department_id', flat=True).distinct()
        )
        depts = list(Department.objects.filter(id__in=managed_ids).order_by('name'))

        # โหลด monthly budget ทั้งหมดของเดือนนี้ใน 1 query
        monthly_budgets = {
            mb.department_id: float(mb.budget)
            for mb in DepartmentMonthlyBudget.objects.filter(year=year, month=mon)
        }

        dept_list = []
        total_budget = 0
        total_used = 0

        for d in depts:
            budget = monthly_budgets.get(d.id, 0)
            reqs = OTRequest.objects.filter(
                staff__department=d,
                work_date__year=year,
                work_date__month=mon,
                status__in=APPROVED_STATUSES,
            )
            used = sum(_calc_ot_amount(r.ot_hours, r.day_type) for r in reqs)
            remaining = budget - used
            pct = round(used / budget * 100) if budget > 0 else 0
            dept_list.append({
                'id': d.id, 'name': d.name, 'code': d.code,
                'ot_budget': budget, 'budget': budget,
                'used': used, 'remaining': remaining, 'pct': pct,
            })
            total_budget += budget
            total_used += used

        total_pct = round(total_used / total_budget * 100) if total_budget > 0 else 0

        # Trend: past 6 months
        today = datetime.date.today()
        trend = []
        for i in range(5, -1, -1):
            m2 = today.month - i
            y2 = today.year
            while m2 <= 0:
                m2 += 12
                y2 -= 1
            reqs2 = OTRequest.objects.filter(
                work_date__year=y2, work_date__month=m2,
                status__in=APPROVED_STATUSES,
            )
            a2 = sum(_calc_ot_amount(r.ot_hours, r.day_type) for r in reqs2)
            trend.append({'m': THAI_MONTHS_SHORT[m2 - 1], 'a': a2})

        # no_ot_depts: dept names without any OT requests this month
        depts_with_ot = set(
            OTRequest.objects.filter(work_date__year=year, work_date__month=mon)
            .values_list('staff__department_id', flat=True).distinct()
        )
        no_ot_dept_names = [d.name for d in depts if d.id not in depts_with_ot]

        return Response({
            'month': month_param,
            'departments': dept_list,
            'total_budget': total_budget,
            'total_used': total_used,
            'total_pct': total_pct,
            'trend': trend,
            'no_ot_depts': no_ot_dept_names,
        })

    # POST — อัปเดต monthly budget
    # payload: {'month': 'YYYY-MM', 'budgets': [{'id': dept_id, 'ot_budget': 50000}, ...]}
    import datetime as _dt
    payload_month = request.data.get('month') if isinstance(request.data, dict) else None
    if not payload_month:
        payload_month = _dt.date.today().strftime('%Y-%m')
    updates = request.data.get('budgets') if isinstance(request.data, dict) else request.data
    if not isinstance(updates, list):
        return Response({'error': 'expected list in budgets field'}, status=400)

    p_year, p_mon = map(int, payload_month.split('-'))
    updated_depts = []
    for item in updates:
        try:
            dept = Department.objects.get(id=item['id'])
            new_budget = item.get('ot_budget', 0)
            DepartmentMonthlyBudget.objects.update_or_create(
                department=dept, year=p_year, month=p_mon,
                defaults={'budget': new_budget, 'set_by': request.user},
            )
            updated_depts.append((dept, new_budget))
        except Department.DoesNotExist:
            pass

    # แจ้งเตือน depthead ทุกคนในแผนกที่อัปเดตงบ
    checker_name = request.user.get_full_name() or request.user.username
    thai_month_names = ['มกราคม','กุมภาพันธ์','มีนาคม','เมษายน','พฤษภาคม','มิถุนายน','กรกฎาคม','สิงหาคม','กันยายน','ตุลาคม','พฤศจิกายน','ธันวาคม']
    month_label = f'{thai_month_names[p_mon - 1]} {p_year + 543}'
    for dept, new_budget in updated_depts:
        depthead_users = User.objects.filter(role='depthead', is_active=True, department=dept)
        for u in depthead_users:
            msg_budget = f'ผู้ตรวจสอบ ({checker_name}) ตั้งงบประมาณ OT แผนก {dept.name} เดือน{month_label} เป็น {float(new_budget):,.0f} บาท'
            notif = Notification.objects.create(
                recipient=u,
                message=msg_budget,
                notif_type='budget_set',
            )
            _push_ws(u.id, {
                'id': notif.id, 'message': notif.message, 'notif_type': notif.notif_type,
                'ot_request': None, 'ot_request_date': None, 'is_read': False,
                'created_at': notif.created_at.isoformat(),
            })
            _send_email(u, f'[SMART OT] {msg_budget[:60]}', msg_budget)

    return Response({'status': 'ok', 'updated': len(updated_depts), 'month': payload_month})


# ─── No-OT Departments (Checker) ──────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def no_ot_departments_view(request):
    import datetime
    month = request.query_params.get('month')  # YYYY-MM
    if not month:
        now = datetime.date.today()
        month = now.strftime('%Y-%m')

    year, mon = map(int, month.split('-'))
    depts_with_ot = OTRequest.objects.filter(
        work_date__year=year,
        work_date__month=mon,
    ).values_list('staff__department_id', flat=True).distinct()

    # เฉพาะแผนกที่มี depthead หรือ deptrep ผูกอยู่ (แผนก "จริง" ในระบบ)
    managed_dept_ids = set(
        User.objects.filter(role__in=['depthead', 'deptrep'], is_active=True)
        .values_list('department_id', flat=True).distinct()
    )
    no_ot = Department.objects.exclude(id__in=depts_with_ot).filter(id__in=managed_dept_ids).order_by('name')
    # แผนกที่แจ้งไม่ประสงค์ส่ง OT เดือนนี้ (NoOTDeclaration)
    declarations = NoOTDeclaration.objects.filter(
        greg_year=year, month=mon,
    ).select_related('department', 'declared_by')

    declared_dept_ids = set(d.department_id for d in declarations)
    data = [{'id': d.id, 'name': d.name, 'code': d.code} for d in no_ot]
    decl_data = [{
        'dept_id': d.department_id,
        'dept_name': d.department.name,
        'declared_by': d.declared_by.get_full_name() or d.declared_by.username,
        'note': d.note,
        'created_at': d.created_at.isoformat(),
    } for d in declarations]

    return Response({'month': month, 'departments': data, 'declarations': decl_data})


# ─── Head Report ──────────────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def head_report_view(request):
    user = request.user
    dept = getattr(user, 'department', None)
    if dept is None:
        return Response({'error': 'ไม่พบแผนก'}, status=400)

    qs = OTRequest.objects.filter(staff__department=dept).order_by('-work_date')

    # filter by month (YYYY-MM)
    month = request.query_params.get('month')
    if month:
        try:
            year, mon = map(int, month.split('-'))
            qs = qs.filter(work_date__year=year, work_date__month=mon)
        except Exception:
            pass

    data = OTRequestSerializer(qs, many=True).data
    total_hours = sum(int(float(r.get('ot_hours') or 0)) for r in data)
    total_amount = sum(
        int(float(r.get('ot_hours') or 0)) * (70 if r.get('day_type') == 'holiday' else 60)
        for r in data
    )
    return Response({
        'department': dept.name,
        'month': month or 'ทั้งหมด',
        'total_requests': len(data),
        'total_hours': total_hours,
        'total_amount': total_amount,
        'requests': data,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def head_report_pdf_view(request):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from django.http import HttpResponse
        import io, os, datetime

        user = request.user
        dept = getattr(user, 'department', None)
        month = request.query_params.get('month', datetime.date.today().strftime('%Y-%m'))

        qs = OTRequest.objects.filter(staff__department=dept)
        if month:
            try:
                year, mon = map(int, month.split('-'))
                qs = qs.filter(work_date__year=year, work_date__month=mon)
            except Exception:
                pass
        qs = qs.order_by('work_date')

        buf = io.BytesIO()
        p = canvas.Canvas(buf, pagesize=A4)
        w, h = A4

        # หา font Thai
        font_name = 'Helvetica'
        for fp in [
            '/usr/share/fonts/truetype/thai/Sarabun-Regular.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        ]:
            if os.path.exists(fp):
                try:
                    pdfmetrics.registerFont(TTFont('Thai', fp))
                    font_name = 'Thai'
                except Exception:
                    pass
                break

        p.setFont(font_name, 16)
        p.drawString(60, h - 60, f'รายงาน OT แผนก {dept.name if dept else ""}')
        p.setFont(font_name, 12)
        p.drawString(60, h - 85, f'เดือน: {month}')

        y = h - 120
        p.setFont(font_name, 10)
        headers = ['ชื่อ', 'วันที่', 'ประเภท', 'ชม.', 'จำนวนเงิน', 'สถานะ']
        x_pos  = [60, 180, 280, 340, 390, 460]
        for i, hdr in enumerate(headers):
            p.drawString(x_pos[i], y, hdr)
        y -= 5
        p.line(60, y, 540, y)
        y -= 15

        total_amt = 0
        for r in qs:
            if y < 60:
                p.showPage()
                y = h - 60
                p.setFont(font_name, 10)
            hrs = int(float(r.ot_hours or 0))
            amt = hrs * (70 if r.day_type == 'holiday' else 60)
            total_amt += amt
            row = [
                r.staff.get_full_name()[:12],
                str(r.work_date),
                'หยุด' if r.day_type == 'holiday' else 'ธรรมดา',
                str(hrs),
                f'{amt:,}',
                r.status,
            ]
            for i, val in enumerate(row):
                p.drawString(x_pos[i], y, val)
            y -= 15

        y -= 10
        p.line(60, y, 540, y)
        y -= 15
        p.setFont(font_name, 11)
        p.drawString(390, y, f'รวม: {total_amt:,} บาท')

        p.save()
        buf.seek(0)
        resp = HttpResponse(buf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="ot_report_{month}.pdf"'
        return resp

    except ImportError:
        return Response({'error': 'reportlab not installed: pip install reportlab'}, status=500)


# -- No-OT Declaration ---

@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def no_ot_declaration_view(request):
    user = request.user

    if request.method == 'GET':
        dept = getattr(user, 'department', None)
        qs = NoOTDeclaration.objects.filter(department=dept).order_by('-greg_year', '-month')[:24]
        data = [{
            'id': d.id,
            'greg_year': d.greg_year,
            'month': d.month,
            'note': d.note,
            'declared_by': d.declared_by.get_full_name(),
            'created_at': d.created_at.isoformat(),
        } for d in qs]
        return Response(data)

    effective_role = get_effective_role(user, request)
    if effective_role != 'depthead':
        return Response({'error': 'depthead only'}, status=403)

    greg_year = request.data.get('greg_year')
    month     = request.data.get('month')
    note      = request.data.get('note', '')

    if not greg_year or not month:
        return Response({'error': 'greg_year and month required'}, status=400)

    dept = getattr(user, 'department', None)
    if not dept:
        return Response({'error': 'no department'}, status=400)

    decl, created = NoOTDeclaration.objects.get_or_create(
        department=dept,
        greg_year=int(greg_year),
        month=int(month),
        defaults={'declared_by': user, 'note': note},
    )
    if not created:
        return Response({'error': 'already declared', 'id': decl.id}, status=409)

    thai_year = int(greg_year) + 543
    month_int = int(month)
    msg = f'[{dept.name}] {user.get_full_name()} declared no OT month {month_int}/{thai_year}'

    deptreps   = User.objects.filter(role='deptrep', is_active=True, department=dept)
    checkers   = User.objects.filter(role='checker', is_active=True)
    recipients = (deptreps | checkers).distinct()

    for recipient in recipients:
        notif = Notification.objects.create(
            recipient=recipient,
            message=msg,
            notif_type='no_ot_declared',
        )
        _push_ws(recipient.id, {
            'id': notif.id, 'message': notif.message, 'notif_type': notif.notif_type,
            'ot_request': None, 'ot_request_date': None, 'is_read': False,
            'created_at': notif.created_at.isoformat(),
        })
        _send_email(recipient, f'[SMART OT] {msg[:60]}', msg)

    return Response({
        'id': decl.id,
        'message': msg,
        'notified': recipients.count(),
    }, status=201)
