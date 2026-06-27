# -*- coding: utf-8 -*-
"""
Management command: import_staff_master
อ่าน Excel รายชื่อบุคลากร 2 ไฟล์ ->สร้าง User + Department + ตำแหน่ง + อีเมล

Usage:
  python manage.py import_staff_master \
      "C:/Users/ASUS/Downloads/04. รายชื่อ + อีเมล.xlsx" \
      --emails "C:/Users/ASUS/Downloads/05.รายชื่อ_อีเมล_tu.ac.th-reg.tu.ac.th.xlsx"

File 04 (primary): ลำดับ | ชื่อ-สกุล | ตำแหน่งงาน | E-mail(@reg)
  - แถวที่ ลำดับ=None + ชื่อ-สกุล มีค่า ->ชื่อแผนก (section header)
  - แถวที่ ลำดับ=ตัวเลข ->พนักงาน

File 05 (optional): ลำดับ | ชื่อ-สกุล | ตำแหน่งงาน | @reg | @tu.ac.th
  - ใช้ mapping @reg ->@tu.ac.th สำหรับ login
"""
import re
from django.core.management.base import BaseCommand
from api.models import User, Department

HONORIFICS = [
    'รองศาสตราจารย์ ดร.', 'ผู้ช่วยศาสตราจารย์ ดร.', 'รองศาสตราจารย์',
    'ศาสตราจารย์ ดร.', 'ศาสตราจารย์', 'ว่าที่ ร.ต.', 'ว่าที่ร.ต.',
    'อาจารย์ ดร.', 'อาจารย์', 'นางสาว', 'นาง', 'นาย', 'ดร.',
]


def strip_honorific(name: str):
    name = name.strip()
    for h in sorted(HONORIFICS, key=len, reverse=True):
        if name.startswith(h):
            name = name[len(h):].strip()
            break
    parts = name.split()
    first = parts[0] if parts else name
    last = ' '.join(parts[1:]) if len(parts) > 1 else ''
    return first, last


def make_username(email_reg: str, email_tu: str, first: str, seq: int) -> str:
    if email_tu:
        return email_tu.split('@')[0]
    if email_reg:
        return email_reg.split('@')[0]
    clean = re.sub(r'[^a-zA-Z0-9฀-๿]', '', first)
    return f'{clean}_{seq}'


class Command(BaseCommand):
    help = 'Import staff master list from Excel into DB (run once at deploy)'

    def add_arguments(self, parser):
        parser.add_argument('file04', help='Path to file 04 (primary staff list)')
        parser.add_argument('--emails', help='Path to file 05 (email mapping @tu.ac.th)', default='')
        parser.add_argument('--password', help='Default password', default='smart2025')
        parser.add_argument('--dry-run', action='store_true', help='Preview without saving')

    def handle(self, *args, **options):
        import openpyxl

        # ── Load email mapping from file 05 ─────────────────────
        tu_email_map = {}  # @reg ->@tu.ac.th
        if options['emails']:
            wb5 = openpyxl.load_workbook(options['emails'], data_only=True)
            ws5 = wb5.active
            for row in ws5.iter_rows(min_row=2, values_only=True):
                vals = list(row) + [None] * 5
                reg_email = str(vals[3] or '').strip().lower()
                tu_email = str(vals[4] or '').strip().lower()
                if reg_email and tu_email:
                    tu_email_map[reg_email] = tu_email
            self.stdout.write(f'Loaded {len(tu_email_map)} email mappings from file 05')

        # ── Parse file 04 ───────────────────────────────────────
        wb4 = openpyxl.load_workbook(options['file04'], data_only=True)
        ws4 = wb4.active

        current_dept = None
        created = 0
        updated = 0
        skipped = 0
        dept_created = 0

        skip_names = {'ชื่อ-สกุล', 'รายชื่อบุคลากรสำนักงานทะเบียนนักศึกษา', ''}

        for row in ws4.iter_rows(min_row=1, values_only=True):
            vals = list(row) + [None] * 4
            seq_raw, name_raw, position_raw, email_raw = vals[:4]

            name_str = str(name_raw or '').strip()
            if not name_str or name_str in skip_names:
                continue

            # Section header (department)
            if seq_raw is None:
                dept_name = name_str
                dept, was_created = Department.objects.get_or_create(
                    name=dept_name,
                    defaults={'code': re.sub(r'[^a-zA-Z0-9฀-๿]', '_', dept_name)[:20].upper()}
                )
                current_dept = dept
                if was_created:
                    dept_created += 1
                    self.stdout.write(f'  [DEPT] สร้างแผนก: {dept_name}')
                continue

            # Employee row
            seq = int(str(seq_raw).strip()) if str(seq_raw).strip().isdigit() else 0
            if seq == 0:
                continue

            first, last = strip_honorific(name_str)
            position = str(position_raw or '').strip().replace('\n', ' ')
            reg_email = str(email_raw or '').strip().lower()
            tu_email = tu_email_map.get(reg_email, '')

            username = make_username(reg_email, tu_email, first, seq)

            if options['dry_run']:
                self.stdout.write(
                    f'  [DRY] #{seq} {first} {last} | dept={current_dept} | '
                    f'user={username} | reg={reg_email} | tu={tu_email} | pos={position[:40]}'
                )
                continue

            user, was_created = User.objects.get_or_create(
                username=username,
                defaults={
                    'first_name': first,
                    'last_name': last,
                    'email': tu_email or reg_email,
                    'notify_email': reg_email,
                    'role': 'staff',
                    'department': current_dept,
                    'is_active': True,
                }
            )

            if was_created:
                user.set_password(options['password'])
                user.save()
                created += 1
                self.stdout.write(f'  [NEW] #{seq} {first} {last} ->{username}')
            else:
                changed = False
                if not user.email and (tu_email or reg_email):
                    user.email = tu_email or reg_email
                    changed = True
                if not user.notify_email and reg_email:
                    user.notify_email = reg_email
                    changed = True
                if user.department is None and current_dept:
                    user.department = current_dept
                    changed = True
                if changed:
                    user.save()
                    updated += 1
                    self.stdout.write(f'  [UPD] #{seq} {first} {last} ->อัปเดตข้อมูล')
                else:
                    skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f'\nเสร็จสิ้น: สร้างใหม่ {created} คน, อัปเดต {updated} คน, '
            f'ข้าม {skipped} คน, แผนกใหม่ {dept_created} แผนก'
        ))
        if options['dry_run']:
            self.stdout.write(self.style.WARNING('(DRY RUN — ไม่ได้บันทึกจริง)'))
